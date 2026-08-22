from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import Workbook

from data_loader import FAMILY_BOEING, parse_workbook


def sample_workbook(model: str = "B747") -> bytes:
    workbook = Workbook()
    workbook.active.title = "工作表1"
    sheet = workbook.create_sheet("评估数据")
    headers = [
        "提交时间", "填写ID", "被评估人姓名", "评估日期", "所属单位", "机型", "技术等级",
        "总飞行时间", "本机型经历时间", "评估员姓名", "单科总得分", "教员双盲训前讲评总得分",
        "决策偏差", "喊话偏差", "单科总得分2", "教员双盲模拟机评估总得分",
    ]
    sheet.append(headers)
    sheet.append([
        "2026-08-21", "id-1", "测试教员", "2026-08-20", "测试单位", "B747", "型别教员",
        8000, 1600, "测试评估员", 20, 95, "无异常（不扣分）", "未喊出中断（-2分）", 18, 98,
    ])
    output = BytesIO()
    workbook.save(output)
    sheet.cell(row=2, column=6).value = model
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class ParserTests(unittest.TestCase):
    def test_skips_blank_sheet_and_extracts_deduction(self):
        summary, ratings, deductions, quality, columns = parse_workbook(sample_workbook(), "波音测试.xlsx")
        self.assertEqual(summary["工作表"], "评估数据")
        self.assertEqual(summary["评估人数"], 1)
        self.assertEqual(ratings.iloc[0]["机型类别"], FAMILY_BOEING)
        self.assertEqual(ratings.iloc[0]["失分"], 2)
        self.assertEqual(ratings.iloc[0]["计算模拟机总分"], 98)
        self.assertEqual(len(deductions), 1)
        self.assertEqual(deductions.iloc[0]["规则状态"], "已识别")
        self.assertFalse(quality.empty)
        self.assertGreater(len(columns), 10)

    def test_unknown_model_is_flagged(self):
        payload = sample_workbook("X999")
        summary, ratings, _, quality, _ = parse_workbook(payload, "未知机型.xlsx")
        self.assertEqual(ratings.iloc[0]["机型类别"], "未识别")
        self.assertIn("未识别机型规则", quality.iloc[0]["问题"])


if __name__ == "__main__":
    unittest.main()
